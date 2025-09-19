# timesheet_app.py — v25.0 (export code taken verbatim from your provided file)
# NOTE: This version keeps the original Excel-based workflow and exports:
# - Per-job: "mm-dd-yyyy - {Job Number} - Daily Time Import.xlsx"
# - Daily:   "mm-dd-yyyy – Daily Time.xlsx" (en dash, matches your template naming)
#
# Place this file next to:
#   - TimeSheet Apps.xlsx
#   - Daily Time.xlsx
#
# If your workbook lives elsewhere, put its full path into a file named
#   timesheet_default_path.txt
# in the same folder as this app (one line, full path to TimeSheet Apps.xlsx).
#
# This is your original app (v25.0) with its exporting code intact.

import os, json, datetime as dt
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font

st.set_page_config(page_title="Timesheet (Excel)", page_icon=":clipboard:", layout="centered")

# ---------- Initialize session keys ----------
for k, v in {
    "whoami_email": "",
    "entered_app": False,
    "is_admin": False,
    "xlsx_path": os.getenv("STREAMLIT_TIMESHEET_XLSX", "TimeSheet Apps.xlsx"),
    "enforce_users": True,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

APP_DIR = Path(__file__).parent

# ---------- Landing page ----------
def show_landing():
    jpgs = sorted(APP_DIR.glob("*.jpg"))
    logo_path = str(jpgs[0]) if jpgs else None

    left, mid, right = st.columns([1, 2, 1])
    with mid:
        if logo_path:
            st.image(logo_path, width=300)
        email = st.text_input("Your work email", st.session_state.get("whoami_email",""), placeholder="name@ptwenergy.com")
        go = st.button("Enter")
        if go:
            st.session_state["whoami_email"] = (email or "").strip()
            st.session_state["entered_app"] = True
            default_xlsx = os.getenv("STREAMLIT_TIMESHEET_XLSX", "TimeSheet Apps.xlsx")
            sidecar = APP_DIR / "timesheet_default_path.txt"
            if sidecar.exists():
                try:
                    p = sidecar.read_text().strip()
                    if p: default_xlsx = p
                except Exception:
                    pass
            st.session_state["xlsx_path"] = default_xlsx
            st.rerun()

if (not st.session_state.get("entered_app", False)) or (not st.session_state.get("whoami_email","").strip()):
    show_landing()
    st.stop()

xlsx_path = st.session_state.get("xlsx_path", "TimeSheet Apps.xlsx")

# ---------- Sidebar ----------
with st.sidebar:
    st.header("Settings")
    st.write("Excel workbook path (close Excel while using the app):")
    st.code(xlsx_path)
    st.checkbox("Restrict access via 'Users' sheet", value=st.session_state.get("enforce_users", True), key="enforce_users")
    st.text_input("Your work email", key="whoami_email")

# ---------- Helpers ----------
def _clean_headers(df: pd.DataFrame) -> pd.DataFrame:
    try: df.columns = [str(c).strip() for c in df.columns]
    except Exception: pass
    return df

def _first(cols, names):
    s = {str(c) for c in cols}
    for n in names:
        if n in s: return n
    return None

def _pad_job_area(v) -> str:
    s = str(v).strip()
    return f"{int(s):03d}" if s.isdigit() else s

def _read_time_data_df(xlsx_file: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(xlsx_file, sheet_name="Time Data"); _clean_headers(df); return df
    except Exception:
        cols = ["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number",
                "RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"]
        return pd.DataFrame(columns=cols)

def _ensure_time_data_headers(xlsx_file: str, headers_to_add: list):
    wb = load_workbook(xlsx_file)
    if "Time Data" not in wb.sheetnames:
        ws = wb.create_sheet("Time Data")
        ws.append(["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number",
                   "RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"])
        wb.save(xlsx_file); return
    ws = wb["Time Data"]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    changed = False
    for h in headers_to_add:
        if h not in headers:
            ws.cell(row=1, column=len(headers)+1, value=h); headers.append(h); changed = True
    if changed: wb.save(xlsx_file)

def _append_dict_row_to_time_data(xlsx_file: str, payload: dict):
    wb = load_workbook(xlsx_file)
    if "Time Data" not in wb.sheetnames:
        ws = wb.create_sheet("Time Data")
        ws.append(["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number",
                   "RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"])
    ws = wb["Time Data"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    if not any(headers):
        headers = ["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number",
                   "RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"]
        for idx, h in enumerate(headers, start=1): ws.cell(row=1, column=idx, value=h)
    row_vals = [payload.get(h, "") for h in headers]
    ws.append(row_vals); wb.save(xlsx_file)

@st.cache_data(ttl=30)
def load_lookups(path: str):
    employees = pd.read_excel(path, sheet_name="Employee List"); _clean_headers(employees)
    jobs      = pd.read_excel(path, sheet_name="Job Numbers"); _clean_headers(jobs)
    costcodes = pd.read_excel(path, sheet_name="Cost Codes");  _clean_headers(costcodes)
    return {
        "employees": employees, "jobs": jobs, "costcodes": costcodes,
        "emp_name_col": _first(employees.columns, ["Employee Name","Name"]),
        "emp_num_col":  _first(employees.columns, ["Person Number","Employee Number","Emp #"]),
        "emp_trade_col":_first(employees.columns, ["Override Trade Class","Trade Class"]),
        "job_num_col":  _first(jobs.columns, ["JOB #","Job Number","Job #"]),
        "job_area_col": _first(jobs.columns, ["AREA #","Job Area","Area #"]),
        "job_desc_col": _first(jobs.columns, ["DESCRIPTION","Area Description","Description","Area Name"]),
        "cost_code_col":_first(costcodes.columns, ["Cost Code","Class Type"]),
        "paycode_map": {"REG":"211","OT":"212","SUBSISTENCE":"261"},
    }

@st.cache_data(ttl=30)
def load_users(path: str):
    try:
        users = pd.read_excel(path, sheet_name="Users"); _clean_headers(users)
    except Exception:
        users = pd.DataFrame(columns=["Email","Active","Type"])
    email_col  = _first(users.columns, ["Email","User Email","Work Email","MAIL"])
    active_col = _first(users.columns, ["Active","ACTIVE"])
    admin_col  = _first(users.columns, ["Type","User Type","Role","ROLE","Is Admin","IsAdmin","ADMIN","IS_ADMIN"])
    return users, email_col, active_col, admin_col

def is_user_admin(users_df: pd.DataFrame, email_col: str, admin_col: str, email: str) -> bool:
    if not email_col or not admin_col: return False
    try:
        row = users_df.loc[users_df[email_col].astype(str).str.strip().str.lower() == str(email).strip().lower()].iloc[0]
        val = str(row.get(admin_col, "")).strip().lower()
        if admin_col.lower() == "type":
            return val == "admin"
        return val in {"true","t","yes","y","1","admin"}
    except Exception:
        return False

# Guard workbook
if not os.path.exists(xlsx_path):
    st.error("Excel workbook not found.\n\nCreate **timesheet_default_path.txt** in the app folder with the full path to your 'TimeSheet Apps.xlsx', or set STREAMLIT_TIMESHEET_XLSX. Make sure the file exists and is not open.")
    st.stop()

# Users gating
users_df, users_email_col, users_active_col, users_admin_col = load_users(xlsx_path)
if st.session_state.get("enforce_users", True):
    allowed = set()
    if users_email_col and users_active_col and users_email_col in users_df and users_active_col in users_df:
        try:
            allowed = set(
                users_df.loc[users_df[users_active_col] == True, users_email_col]
                .dropna().astype(str).str.strip().str.lower()
            )
        except Exception:
            allowed = set()
    if users_email_col and st.session_state.get("whoami_email","").strip().lower() not in allowed:
        st.error("You're not on the 'Users' sheet as Active. Ask an admin to add you.")
        st.stop()

# compute admin flag
st.session_state["is_admin"] = is_user_admin(users_df, users_email_col, users_admin_col, st.session_state.get("whoami_email",""))

# Load lookups
look = load_lookups(xlsx_path)
employees = look["employees"]; jobs = look["jobs"]; costcodes = look["costcodes"]
emp_name_col=look["emp_name_col"]; emp_num_col=look["emp_num_col"]; emp_trade_col=look["emp_trade_col"]
job_num_col=look["job_num_col"]; job_area_col=look["job_area_col"]; job_desc_col=look["job_desc_col"]
cost_code_col=look["cost_code_col"]; paycode_map=look["paycode_map"]

# ---------- Entry UI ----------
st.subheader("Timesheet Entry")
date_val = st.date_input("Date", dt.date.today())

emp_opts = employees[emp_name_col].astype(str).tolist() if emp_name_col else []
sel_emps = st.multiselect("Employees", emp_opts)

job_opts = jobs[job_num_col].astype(str).unique().tolist() if job_num_col else []
sel_job  = st.selectbox("Job Number", [""] + job_opts)

area_labels, area_map = [], {}
if sel_job and job_area_col:
    df = jobs.loc[jobs[job_num_col].astype(str)==str(sel_job)].copy(); _clean_headers(df)
    for _, r in df.iterrows():
        code = _pad_job_area(r.get(job_area_col, ""))
        desc = str(r.get(job_desc_col,"") or "").strip()
        lab = f"{code} - {desc}" if desc else code
        if lab not in area_map: area_labels.append(lab); area_map[lab]=code
sel_area_label = st.selectbox("Job Area", [""] + area_labels)
sel_area_code = area_map.get(sel_area_label, "")

# Cost codes -> Active only
def _only_active_costcodes(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy(); _clean_headers(df2)
    active_col = _first(df2.columns, ["Active","Is Active","Enabled","ACTIVE","IS ACTIVE","ENABLED"])
    if active_col:
        def _truthy(x):
            if isinstance(x, bool): return x
            s = str(x).strip().lower()
            return s in {"true","t","yes","y","1","active","enabled"}
        return df2[df2[active_col].apply(_truthy)]
    status_col = _first(df2.columns, ["Status","STATUS"])
    if status_col:
        return df2[df2[status_col].astype(str).str.strip().str.lower() == "active"]
    end_col = _first(df2.columns, ["End Date","Inactive Date","Date End","END DATE"])
    if end_col:
        return df2[(df2[end_col].isna()) | (df2[end_col].astype(str).str.strip() == "")]
    return df2

active_costcodes = _only_active_costcodes(costcodes)

def build_cost_labels(df, code_col):
    df2=df.copy(); _clean_headers(df2)
    desc_col = _first(df2.columns, ["Cost Code Description","Class Type Description","Description","Cost Code Name","Name"])
    labels, mapping = [], {}
    for _, r in df2.iterrows():
        code = str(r.get(code_col,"") or "").strip()
        if not code: continue
        desc = str(r.get(desc_col,"") or "").strip() if desc_col else ""
        lab = f"{code} - {desc}" if desc else code
        if lab not in mapping: labels.append(lab); mapping[lab]=code
    return labels, mapping

code_labels, code_map = build_cost_labels(active_costcodes, cost_code_col)
sel_code_label = st.selectbox("Class Type (Cost Code)", [""] + code_labels)
sel_code_code = code_map.get(sel_code_label, "")

rt_hours = st.number_input("RT Hours (per employee)", min_value=0.0, max_value=24.0, step=0.5, value=0.0)
ot_hours = st.number_input("OT Hours (per employee)", min_value=0.0, max_value=24.0, step=0.5, value=0.0)
desc     = st.text_area("Comments (optional)", "", height=100)

if st.button("Submit"):
    if not sel_emps:
        st.warning("Select at least one employee.")
    elif not sel_job or not sel_area_code or not sel_code_code:
        st.warning("Select Job, Area, and Class Type.")
    else:
        try: _ensure_time_data_headers(xlsx_path, ["RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"])
        except Exception: pass
        successes = 0
        for emp_name in sel_emps:
            try:
                emp_row = employees.loc[employees[emp_name_col].astype(str) == str(emp_name)].iloc[0]
            except Exception:
                st.error(f"Employee '{emp_name}' not found."); continue
            payload = {
                "Job Number": str(sel_job),
                "Job Area": _pad_job_area(sel_area_code),
                "Date": date_val.strftime("%Y-%m-%d"),
                "Name": emp_name,
                "Class Type": sel_code_code,
                "Trade Class": emp_row.get(emp_trade_col,""),
                "Employee Number": emp_row.get(emp_num_col,""),
                "RT Hours": float(rt_hours),
                "OT Hours": float(ot_hours),
                "Night Shift": "",
                "Premium Rate / Subsistence Rate / Travel Rate": "",
                "Comments": desc,
            }
            try:
                _append_dict_row_to_time_data(xlsx_path, payload); successes += 1
            except Exception as e:
                st.error(f"Failed to append row for {emp_name}: {e}")
        if successes: st.success(f"Added {successes} row(s) to 'Time Data'.")

# ---------- What's been added for this day ----------
st.markdown("---")
st.subheader("What's been added for this day")

filter_by_job = st.checkbox("Filter by selected Job Number", value=False)

td_all = _read_time_data_df(xlsx_path)
if not td_all.empty:
    td_all["__DateStr"] = td_all["Date"].astype(str).str[:10]
    mask = td_all["__DateStr"] == date_val.strftime("%Y-%m-%d")
    if filter_by_job and sel_job:
        mask = mask & (td_all["Job Number"].astype(str).str.strip() == str(sel_job))
    day_df = td_all[mask].copy()
else:
    day_df = pd.DataFrame(columns=["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number",
                                   "RT Hours","OT Hours","Night Shift","Premium Rate / Subsistence Rate / Travel Rate","Comments"])

# ---- Robust mapping from displayed rows to Excel rows ----
def _norm_val(v):
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    try:
        f = float(v)
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
        return ("%.2f" % f).rstrip("0").rstrip(".")
    except Exception:
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                d = datetime.strptime(s, fmt).date()
                return d.strftime("%Y-%m-%d")
            except Exception:
                pass
        return s

def map_rows_to_excel_indices(xlsx_file: str, filt_df: pd.DataFrame) -> pd.DataFrame:
    try:
        wb = load_workbook(xlsx_file, read_only=True, data_only=True)
        if "Time Data" not in wb.sheetnames:
            return filt_df.assign(ExcelRow=None)
        ws = wb["Time Data"]
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        excel_rec = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rec = {headers[i]: _norm_val(row[i]) for i in range(min(len(headers), len(row)))}
            excel_rec.append((r_idx, rec))
        key_cols = [c for c in ["Job Number","Job Area","Date","Name","Class Type","Employee Number","RT Hours","OT Hours","Comments"] if c in filt_df.columns]
        excel_rows = []
        for _, r in filt_df.iterrows():
            target = {c: _norm_val(r.get(c, "")) for c in key_cols}
            match_row = None
            for ridx, rec in excel_rec:
                ok = True
                for c, val in target.items():
                    if val == "":
                        continue
                    if rec.get(c, "") != val:
                        ok = False; break
                if ok:
                    match_row = ridx
                    break
            excel_rows.append(match_row)
        return filt_df.assign(ExcelRow=excel_rows)
    except Exception:
        return filt_df.assign(ExcelRow=None)

day_df = map_rows_to_excel_indices(xlsx_path, day_df)

if day_df.empty:
    st.caption("empty")
else:
    show_cols = ["Job Number","Job Area","Date","Name","Class Type","Trade Class","Employee Number","RT Hours","OT Hours","Comments"]
    show_cols = [c for c in show_cols if c in day_df.columns]
    display_df = day_df.reset_index(drop=True).copy()
    display_df.insert(0, "IDX", display_df.index)
    st.dataframe(display_df[["IDX"] + show_cols], use_container_width=True, hide_index=True)

    st.write("Select rows to delete")
    base_options = [f"{i} — {r.get('Name','')} ({r.get('Job Number','')}/{_pad_job_area(r.get('Job Area',''))})"
                    for i, r in display_df.iterrows()]
    options = ["ALL — Delete all rows shown"] + base_options
    to_del = st.multiselect("Choose options", options, placeholder="Choose options")

    if st.button("Delete selected rows"):
        delete_all = any(opt.startswith("ALL") for opt in to_del)
        if delete_all:
            excel_rows = [int(r) for r in display_df["ExcelRow"].dropna().astype(int).tolist()]
        else:
            idxs = []
            for opt in to_del:
                if opt.startswith("ALL"): 
                    continue
                try:
                    idx = int(opt.split("—")[0].strip())
                    idxs.append(idx)
                except Exception:
                    pass
            excel_rows = [int(display_df.loc[i, "ExcelRow"]) for i in idxs if pd.notna(display_df.loc[i, "ExcelRow"])]
        if not excel_rows:
            st.warning("Nothing matched to delete.")
        else:
            try:
                wb = load_workbook(xlsx_path)
                ws = wb["Time Data"]
                for r in sorted(set(excel_rows), reverse=True):
                    ws.delete_rows(r, 1)
                wb.save(xlsx_path)
                st.success(f"Deleted {len(set(excel_rows))} row(s). Refreshing…")
                st.rerun()
            except Exception as e:
                st.error(f"Failed to delete rows: {e}")

# ---------- Export helpers & UI (Admin only) ----------
def build_export_rows(td_subset: pd.DataFrame, employee_list: pd.DataFrame, paycode_map: Dict[str,str]) -> pd.DataFrame:
    td = td_subset.copy(); _clean_headers(td)
    el = employee_list.copy(); _clean_headers(el)
    rows = []
    for _, r in td.iterrows():
        reg_h = float(r.get("RT Hours",0) or 0.0)
        ot_h  = float(r.get("OT Hours",0) or 0.0)
        base = {
            "Date": pd.to_datetime(r.get("Date","")).strftime("%Y-%m-%d"),
            "Time Record Type": "", "Person Number": r.get("Employee Number",""),
            "Employee Name": r.get("Name",""), "Override Trade Class": r.get("Trade Class",""),
            "Post To Payroll": "Y", "Cost Code / Phase": r.get("Class Type",""),
            "JobArea": _pad_job_area(r.get("Job Area","")), "Scope Change": "", "Pay Code": "", "Hours": 0.0,
            "Night Shift": "", "Premium Rate / Subsistence Rate / Travel Rate": r.get("Premium Rate / Subsistence Rate / Travel Rate",""),
            "Comments": "",
        }
        if reg_h>0:
            t=base.copy(); t["Pay Code"]=paycode_map.get("REG","211"); t["Hours"]=reg_h; rows.append(t)
        if ot_h>0:
            t=base.copy(); t["Pay Code"]=paycode_map.get("OT","212");  t["Hours"]=ot_h; rows.append(t)
    HEADERS = ['Date','Time Record Type','Person Number','Employee Name','Override Trade Class','Post To Payroll','Cost Code / Phase','JobArea','Scope Change','Pay Code','Hours','Night Shift','Premium Rate / Subsistence Rate / Travel Rate','Comments']
    return pd.DataFrame(rows, columns=HEADERS)

def write_formatted(out_df: pd.DataFrame, out_path: Path):
    import xlsxwriter
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(out_path); ws = wb.add_worksheet("TimeEntries")
    HEADERS = list(out_df.columns)
    base_header = wb.add_format({"bold": True, "bg_color": "#4F81BD", "font_color":"#FFFFFF", "border":1, "align":"center"})
    text_fmt = wb.add_format({"border":1, "align":"center"})
    num_fmt  = wb.add_format({"border":1, "align":"center", "num_format":"0.00"})
    date_fmt = wb.add_format({"border":1, "align":"center", "num_format":"yyyy-mm-dd"})
    for c,h in enumerate(HEADERS): ws.write(0,c,h,base_header)
    for r in range(len(out_df)):
        row=out_df.iloc[r]
        for c,h in enumerate(HEADERS):
            val=row[h]; fmt=num_fmt if h=="Hours" else (date_fmt if h=="Date" else text_fmt)
            ws.write(1+r,c,val if pd.notna(val) else "", fmt)
    wb.close()

def export_all_jobs_for_date(xlsx: str, date_str: str, outdir: str, user: str, paycode_map: Dict[str,str]):
    td = _read_time_data_df(xlsx); el = pd.read_excel(xlsx, sheet_name="Employee List")
    _clean_headers(el); _clean_headers(td)
    if "Date" not in td.columns:
        st.warning("No 'Date' column in Time Data."); return []
    date_col = td["Date"].astype(str).str[:10]
    subset = td[date_col == date_str].copy()
    if subset.empty:
        st.warning(f"No time entries found for {date_str}"); return []
    jobs = sorted(subset["Job Number"].astype(str).str.strip().unique().tolist()) if "Job Number" in subset.columns else ["ALL"]
    out_paths=[]; dt_obj=pd.to_datetime(date_str).date(); month=dt_obj.strftime("%B")
    wb = load_workbook(xlsx)
    if "Exports Log" not in wb.sheetnames:
        ws_log = wb.create_sheet("Exports Log")
        ws_log.append(["LogID","Date","Job Number","Entries Count","File Name","OneDrive Path","Share Link","Triggered By","Triggered At","Status","Notes"])
    ws_log = wb["Exports Log"]
    for job in jobs:
        sub = subset if job=="ALL" else subset[subset["Job Number"].astype(str).str.strip()==job]
        out_df = build_export_rows(sub, el, paycode_map)
        file_name = f"{dt_obj.strftime('%m-%d-%Y')} - {job if job!='ALL' else 'ALL_JOBS'} - Daily Time Import.xlsx"
        out_path = Path(outdir)/month/file_name
        write_formatted(out_df, out_path); out_paths.append(out_path)
        ws_log.append([f"{date_str}-{job}-{datetime.utcnow().strftime('%H%M%S')}", date_str, job, len(out_df), out_path.name, str(out_path.parent), "", user, datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"), "Created", ""])
    wb.save(xlsx); return out_paths

def export_daily_time_report(xlsx: str, template_path: str, date_str: str, outdir: str, user: str=""):
    if not os.path.exists(template_path): return None
    td = _read_time_data_df(xlsx); _clean_headers(td)
    if "Date" not in td.columns: return None
    dmask = td["Date"].astype(str).str[:10]==date_str
    day = td[dmask].copy()
    if day.empty: return None
    try:
        wb = load_workbook(template_path); ws = wb.active
        date_obj = pd.to_datetime(date_str).date()
        ws.cell(row=5, column=7, value=date_obj)  # G5
        outdir_p = Path(outdir)/date_obj.strftime("%B"); outdir_p.mkdir(parents=True, exist_ok=True)
        out_name = f"{date_obj.strftime('%m-%d-%Y')} – Daily Time.xlsx"
        out_path = outdir_p/out_name
        wb.save(out_path)
        logwb = load_workbook(xlsx)
        if "Exports Log" not in logwb.sheetnames:
            wslog = logwb.create_sheet("Exports Log")
            wslog.append(["LogID","Date","Job Number","Entries Count","File Name","OneDrive Path","Share Link","Triggered By","Triggered At","Status","Notes"])
        wslog = logwb["Exports Log"]
        wslog.append([f"{date_str}-DAILY-{datetime.utcnow().strftime('%H%M%S')}", date_str, "ALL", len(day), out_path.name, str(outdir_p), "", user, datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"), "Created", ""])
        logwb.save(xlsx)
        return str(out_path)
    except Exception:
        return None

# ---------- Export UI (Admin only) ----------
if st.session_state.get("is_admin", False):
    st.markdown("---")
    st.subheader("Export Day → TimeEntries (numeric pay codes, padded JobArea)")
    user = st.text_input("Triggered by (optional)", st.session_state.get('whoami_email', ''))
    default_outdir = str(Path(xlsx_path).parent)
    try:
        with open("user_defaults.json","r") as f:
            d=json.load(f); default_outdir=d.get(user, default_outdir)
    except Exception: pass
    outdir_input = st.text_input("Export Directory", value=default_outdir, key="export_outdir")
    if st.button("Set to Default"): 
        try:
            with open("user_defaults.json","r") as f: d=json.load(f)
        except Exception: d={}
        d[user]=outdir_input
        with open("user_defaults.json","w") as f: json.dump(d,f)
        st.success("Export directory set as default.")
    with st.form("export_form"):
        export_date = st.date_input("Export Date", dt.date.today())
        do_export = st.form_submit_button("Export ALL Jobs for Date")
        if do_export:
            outdir = st.session_state.get("export_outdir", str(Path(xlsx_path).parent))
            try:
                paths = export_all_jobs_for_date(
                    xlsx=xlsx_path,
                    date_str=export_date.strftime("%Y-%m-%d"),
                    outdir=outdir,
                    user=user,
                    paycode_map=look["paycode_map"],
                )
                daily_path = export_daily_time_report(
                    xlsx=xlsx_path,
                    template_path=str(APP_DIR / "Daily Time.xlsx"),
                    date_str=export_date.strftime("%Y-%m-%d"),
                    outdir=outdir,
                    user=user,
                )
                n_daily = 1 if daily_path else 0
                if len(paths)+n_daily==0: st.warning("No matching rows for that date. No files created.")
                else: st.success(f"Created {len(paths)} TimeEntries file(s) and {n_daily} Daily Time report(s).")
            except Exception as e:
                st.error(f"Export failed: {e}")
else:
    st.info("You’re signed in, but your account isn’t marked Admin on the 'Users' sheet. Ask an admin to set your role to Admin if you need access to exports.")

st.caption("Exports default to your workbook's folder. You can change the Export Directory to a folder you own (e.g., Documents or a mapped SharePoint drive).")
